import qrcode
import qrcode.image.svg

from openpyxl import load_workbook
from datetime import datetime, date, timedelta
from collections import defaultdict
from jinja2 import Environment, FileSystemLoader

def main():

    # Load data
    workbook = load_workbook('data.xlsx')
    trees = load_data(workbook, 'trees')
    yields = load_data(workbook, 'yields')
    weather = load_data(workbook, 'weather')

    # Load templates
    environment = Environment(loader=FileSystemLoader("templates/"))
    
    environment.filters['format_date'] = format_date
    environment.filters['format_latitude'] = format_latitude
    environment.filters['format_longitude'] = format_longitude

    template_tree = environment.get_template("tree.html")
    template_inventory = environment.get_template("inventory.html")
    template_season = environment.get_template("season.html")

    # Iterate over trees
    for tree in trees:

        # Generate tree webpage
        page = template_tree.render(
            tree=tree,
            yields=get_yields(yields, tree)
        )
        render_page(
            page, 
            filename=f"docs/tree/{tree['id']}.html"
        )

        # Generate tree QR code
        img = qrcode.make(
            data=f'samela.io/quaboag-maple-co/tree/{tree["id"]}',
            image_factory=qrcode.image.svg.SvgPathImage,
            border=0
        )
        img.save(f'docs/assets/qr/{tree["id"]}.svg')

    # Iterate over seasons
    seasons = get_seasons(yields)

    for season in seasons:

        # Ignore future seasons that haven't happened yet
        if season > datetime.now().year:
            continue

        # Get source data
        season_trees = get_yields_season(yields, season)
        season_weather = get_season_weather(season, weather)


        # Iterate over season tress to count metrics
        season_start_dates = []
        season_end_dates = []
        total_sap = 0
        for t in season_trees:
            total_sap+=t['yield']
            season_start_dates.append(t['start'])
            season_end_dates.append(t['end'])

        season_start = min(season_start_dates).date()
        season_end = max(season_end_dates).date()
        trees_tapped = len(season_trees)
        season_length = (season_end-season_start).days
        sap_per_tree = round(total_sap/trees_tapped, 2)

        # Iterate over days in season to format weather data
        start = date(season, 1, 1)
        end = date(season, 4, 2)
        days = []
        temps = []
        colors = []
        cold_days = 0
        warm_days = 0
        good_days = 0
        while start < end:
            # Write day
            days.append(start.isoformat())

            # Write temps
            t = season_weather[start]
            temp_max = t['high']
            temp_min = t['low']
            temps.append([temp_min, temp_max])

            # Write colors & count by type
            if start < season_start or start > season_end or temp_max == '-':
                color = 'gray'
            elif temp_min <= 32 and temp_max <= 32:
                # Cold
                color = 'blue'
                cold_days += 1
            elif temp_min > 32 and temp_max > 32:
                # Warm
                color = 'red'
                warm_days += 1
            elif temp_min <= 32 and temp_max > 32:
                # Good
                color = 'green'
                good_days += 1

            colors.append(color)

            start += timedelta(days=1)

        # Format data into dict for page
        data = {
            'season': season,
            'season_start': season_start,
            'season_end': season_end,
            'season_length': season_length,
            'trees_tapped': trees_tapped,
            'total_sap': total_sap,
            'sap_per_tree': sap_per_tree,
            'trees': season_trees,
            'warm_days': warm_days,
            'cold_days': cold_days,
            'good_days': good_days,
            'weather': {
                'days': days,
                'temps': temps,
                'colors': colors
            }
        }

        # Render season pages
        page = template_season.render(data=data)
        render_page(
            page, 
            filename=f"docs/season/{season}.html"
        )

    # Render inventory page
    page = template_inventory.render(trees=trees)
    render_page(page, "docs/index.html")


def render_page(page, filename):
    with open(filename, mode="w", encoding="utf-8") as message:
        message.write(page)
        print(f"... wrote {filename}")

def get_seasons(yields):
    seasons = []
    for y in yields:
        s = y['season']
        if not s in seasons:
            seasons.append(s)
    return seasons

def get_season_weather(season, weather):

    season_weather = {}
    for w in weather:
        if w['season'] == season:
            season_weather[w['day'].date()] = w
    return season_weather

def get_yields_season(yields, season):
    data = []
    for y in yields:
        if y['season'] == season:
            data.append(y)
    return data

def get_yields(yields, tree):
    data = []
    for y in yields:
        if y['tree'] == tree['id']:
            if isinstance(y['end'], datetime):
                y['days'] = (y['end']-y['start']).days
            else:
                y['days'] = '-'
            data.append(y)
    return data

def load_data(workbook, worksheet):
    first = True
    headers = None
    data = []
    for i, row in enumerate(workbook[worksheet].iter_rows(values_only=True)):
        if first:
            headers = row
            first = False
        else:
            rowdict = dict()
            for i, header in enumerate(headers):
                if header in ['id', 'tree']:
                    rowdict[header] = f"{row[i]:02}"
                else:
                    rowdict[header] = row[i]
            data.append(rowdict)
    return data

def format_date(d):
    if isinstance(d, datetime) or isinstance(d, date):
        return d.strftime('%b %-d')
    else:
        return d

def format_latitude(coordinate):
    degrees, minute, second = decdeg2dms(coordinate)
    
    
    if degrees < 0:
        direction = 'S'
    else:
        direction = 'N'

    return f'{round(abs(degrees))}°{round(minute)}\'{round(second)}" {direction}'

def format_longitude(coordinate):
    degrees, minute, second = decdeg2dms(coordinate)
    
    if degrees < 0:
        direction = 'W'
    else:
        direction = 'E'

    return f'{round(abs(degrees))}°{round(minute)}\'{round(second)}" {direction}'

def decdeg2dms(dd):
   is_positive = dd >= 0
   dd = abs(dd)
   minutes,seconds = divmod(dd*3600,60)
   degrees,minutes = divmod(minutes,60)
   degrees = degrees if is_positive else -degrees
   return (degrees,minutes,seconds)

if __name__ == '__main__':
    main()
